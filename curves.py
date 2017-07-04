# pip install scipy openpyxl numpy matplotlib
# python curves.py <file.xlsx>

import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import numpy as np

SCURVES_NB = 124
VCAL_RANGE = 34

# Logistic function
def logistic(x, A, B):
    return 100. / (1. + np.exp(-1. * A * (x - B)))



# Load the XLSX document
if len(sys.argv) < 2:
    print "Usage: python curves.py <file>"
    sys.exit()

wb = load_workbook(sys.argv[1])
wb.active = 1
ws = wb.active

# Extract the data
DACs = None
efficiency = [ ]
for i in range(1, SCURVES_NB+2):
    data = [ ]
    for j in range(1, VCAL_RANGE):
        col = get_column_letter(i)
        label = col + str(j)
        data.append(ws[label].value)
    if i == 1:
        DACs = np.array(data)
    else:
        efficiency.append(np.array(data))

# Fit lines
sigmas = [ ]
mus = [ ]
for data in efficiency:
    try:
        popt, pcov = curve_fit(logistic, DACs, data)
        sigmas.append(popt[0])
        mus.append(popt[1])
    except RuntimeError:
        print "Fit failed"


plt.suptitle('SCurves')
plt.subplot(2, 2, 1)
for points in efficiency:
    plt.plot(range(1, VCAL_RANGE), points, '.-')
plt.xlabel('Injected charge [DAC]')
plt.ylabel('Efficiency [%]')

plt.subplot(2, 2, 3)
plt.hist(sigmas, 50)
plt.xlabel('SCurve sigma [DAC]')
plt.ylabel('#')

plt.subplot(2, 2, 4)
plt.hist(mus, 50)
plt.xlabel('SCurve mean [DAC]')
plt.ylabel('#')

plt.show()
